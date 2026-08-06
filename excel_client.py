# -*- coding: utf-8 -*-
"""
excel_client.py
================
Camada de acesso a dados da aplicação REDEB2B — versão simplificada, que
opera exclusivamente sobre um arquivo Excel local (.xlsx), usando
pandas + openpyxl.

Expõe a classe `DataClient` com uma interface de CRUD:
  list_items, get_item, create_item, update_item, delete_item,
  dashboard_aggregates, items_by_date
"""

import os
import logging
import re
import tempfile
import threading
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd

logger = logging.getLogger("redeb2b.excel_client")

# Trava para evitar leituras/escritas simultâneas no mesmo arquivo Excel
# quando o Flask atende mais de uma requisição ao mesmo tempo.
EXCEL_LOCK = threading.RLock()

# Colunas oficiais da planilha REDEB2B, na ordem definida no escopo.
FIELDS = [
    "IDCLIENTE", "CLIENTE", "ENDERECO", "CIDADE", "PRODUTO", "ATIVIDADE",
    "TECNOLOGIA", "VT", "DATADISPARO", "RETORNOPCC", "DATAAGENDAMENTO",
    "DATACONCLUSAO", "OBSERVACAO", "STATUS", "EXECUTADOPOR", "TIPOCABO",
    "METRAGEM", "OBSERVACAOCONCLUSAO", "NUMDRAFT", "ROTA", "USUARIO",
]

# Campos de data que precisam de tratamento especial (ISO yyyy-mm-dd).
DATE_FIELDS = {"DATADISPARO", "DATAAGENDAMENTO", "DATACONCLUSAO"}

# Padrões aceitos ao LER datas do Excel: ISO (yyyy-mm-dd, com ou sem hora) e
# pt-BR (dd/mm/yyyy) — este último é comum quando alguém edita a data
# diretamente na planilha, com o Excel configurado em português.
_ISO_DATE_RE = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})")
_BR_DATE_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})")

MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def _parse_metragem(valor) -> float:
    """Converte o valor de METRAGEM (que pode vir como texto, com vírgula
    decimal, vazio, etc.) em número. Retorna 0.0 se não for possível
    converter."""
    if valor is None:
        return 0.0
    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = texto.replace(",", ".")
    try:
        return float(texto)
    except (ValueError, TypeError):
        return 0.0


class DataClientError(Exception):
    """Erro de negócio/infra do DataClient — sempre com uma mensagem amigável."""

    def __init__(self, message, status_code=400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code


def _strip_accents(value: str) -> str:
    """Remove acentos para permitir buscas 'case/acento-insensitive'."""
    if value is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(value))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()


def _parse_date(value):
    """Converte valores variados (str ISO, str dd/mm/aaaa, datetime, Timestamp
    do pandas) em uma data ISO 'yyyy-mm-dd' (string) para permitir comparação
    e ordenação consistentes, independentemente de como a data foi digitada
    na planilha. Retorna None se o valor estiver vazio ou não reconhecível."""
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")

    s = str(value).strip()
    if not s:
        return None

    m = _ISO_DATE_RE.match(s)
    if m:
        year, month, day = m.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    m = _BR_DATE_RE.match(s)
    if m:
        day, month, year = m.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    # Formato não reconhecido — retorna como veio, truncado, para não quebrar
    # o restante do fluxo (mas não vai comparar corretamente em filtros de data).
    return s[:10]


def _obter_pastas_onedrive() -> list[Path]:
    """Lista as pastas locais do OneDrive candidatas (pessoal e/ou
    corporativo), a partir das variáveis de ambiente que o próprio OneDrive
    define e, como reforço, procurando por pastas 'OneDrive*' na home do
    usuário."""
    candidatos: list[Path] = []

    for variavel in ("OneDriveCommercial", "OneDriveConsumer", "OneDrive"):
        valor = os.environ.get(variavel)
        if valor:
            caminho = Path(valor).expanduser()
            if caminho.exists():
                candidatos.append(caminho)

    for pasta in Path.home().glob("OneDrive*"):
        if pasta.is_dir():
            candidatos.append(pasta)

    resultado: list[Path] = []
    vistos: set[str] = set()
    for pasta in candidatos:
        try:
            chave = str(pasta.resolve()).casefold()
        except OSError:
            chave = str(pasta).casefold()
        if chave not in vistos:
            vistos.add(chave)
            resultado.append(pasta)
    return resultado


def _localizar_excel_no_onedrive(nome_arquivo: str):
    """Procura nome_arquivo dentro das pastas do OneDrive: primeiro em
    alguns locais comuns (raiz, Documentos), depois com busca recursiva
    completa como último recurso. Retorna (caminho_encontrado_ou_None,
    lista_de_pastas_verificadas)."""
    pastas_onedrive = _obter_pastas_onedrive()
    if not pastas_onedrive:
        return None, []

    caminhos_relativos = [
        Path(nome_arquivo),
        Path("Documentos") / nome_arquivo,
        Path("Documents") / nome_arquivo,
    ]

    for pasta in pastas_onedrive:
        for relativo in caminhos_relativos:
            caminho = pasta / relativo
            if caminho.is_file() and not caminho.name.startswith("~$"):
                return caminho, pastas_onedrive

    encontrados: list[Path] = []
    for pasta in pastas_onedrive:
        try:
            encontrados.extend(
                caminho
                for caminho in pasta.rglob(nome_arquivo)
                if caminho.is_file() and not caminho.name.startswith("~$")
            )
        except (PermissionError, OSError):
            continue

    if encontrados:
        encontrados.sort(key=lambda caminho: (len(caminho.parts), str(caminho).casefold()))
        return encontrados[0], pastas_onedrive

    return None, pastas_onedrive


class DataClient:
    """Cliente de dados baseado exclusivamente em um arquivo Excel local."""

    # Cache de classe: guarda o caminho já localizado automaticamente, para
    # não precisar varrer o OneDrive inteiro a cada requisição.
    _caminho_cache: str | None = None

    def __init__(self):
        # Se EXCEL_PATH estiver preenchido no .env, ele tem prioridade e é
        # usado exatamente como configurado (comportamento manual, como antes).
        self.excel_path_env = os.getenv("EXCEL_PATH", "").strip()
        # Se EXCEL_PATH estiver vazio, o app procura automaticamente por este
        # nome de arquivo dentro das pastas do OneDrive sincronizadas nesta
        # máquina (pessoal e/ou corporativo).
        self.excel_filename = os.getenv("EXCEL_FILENAME", "REDE_B2B.xlsx").strip()
        self.excel_sheet = os.getenv("EXCEL_SHEET_NAME", "REDEB2B").strip()

    def _resolver_caminho_excel(self) -> str:
        """Decide qual arquivo usar: 1) EXCEL_PATH explícito no .env, se
        existir; 2) um caminho já localizado automaticamente antes; 3) uma
        nova busca automática pelas pastas do OneDrive."""
        if self.excel_path_env:
            caminho = Path(self.excel_path_env).expanduser()
            if caminho.is_file():
                return str(caminho)
            raise DataClientError(
                f"O caminho configurado em EXCEL_PATH não foi encontrado: '{caminho}'. "
                "Verifique se o caminho está correto, ou apague o valor de EXCEL_PATH "
                "no .env para que o sistema tente localizar o arquivo automaticamente "
                "no OneDrive.",
                status_code=500,
            )

        if DataClient._caminho_cache and Path(DataClient._caminho_cache).is_file():
            return DataClient._caminho_cache

        encontrado, pastas_verificadas = _localizar_excel_no_onedrive(self.excel_filename)
        if encontrado:
            DataClient._caminho_cache = str(encontrado)
            return DataClient._caminho_cache

        pastas_texto = "\n".join(f"- {p}" for p in pastas_verificadas) or "(nenhuma pasta do OneDrive foi encontrada nesta máquina)"
        raise DataClientError(
            f"O arquivo '{self.excel_filename}' não foi encontrado automaticamente no "
            f"OneDrive.\n\nPastas verificadas:\n{pastas_texto}\n\n"
            "Confirme se o OneDrive está instalado, logado e sincronizado, e se o "
            "arquivo está marcado como 'Sempre manter neste dispositivo'. Como "
            "alternativa, defina o caminho manualmente em EXCEL_PATH no .env.",
            status_code=500,
        )

    # ------------------------------------------------------------------
    # Métodos públicos
    # ------------------------------------------------------------------
    def list_items(self, filters=None, page=1, page_size=20, sort=None):
        """Lista itens com filtros, paginação e ordenação.

        filters: dict com chaves opcionais: cliente, id, cidade,
                 executadopor, status, data_inicio, data_fim, q
        sort: string no formato "campo:asc" ou "campo:desc"
        Retorna: dict {items: [...], total: int, page: int, page_size: int}
        """
        filters = filters or {}
        rows = self._excel_read_all()
        rows = self._apply_filters(rows, filters)
        rows = self._apply_sort(rows, sort)

        total = len(rows)
        start = (page - 1) * page_size
        end = start + page_size
        page_items = rows[start:end]

        return {
            "items": page_items,
            "total": total,
            "page": page,
            "page_size": page_size,
        }

    def export_items(self, filters=None, sort=None):
        """Retorna TODOS os itens que atendem aos filtros e à ordenação
        informados, sem paginação — usado para exportar para Excel
        exatamente o que está filtrado na tela."""
        filters = filters or {}
        rows = self._excel_read_all()
        rows = self._apply_filters(rows, filters)
        rows = self._apply_sort(rows, sort)
        return rows

    def analytics(self, ano=None, mes=None):
        """Agregações para a página de Análises (/analises), com uma janela
        de tempo configurável:
          - ano E mes informados -> apenas aquele mês específico
          - só ano informado -> os 12 meses daquele ano
          - nenhum dos dois -> últimos 6 meses (incluindo o mês atual)

        O mês de referência de cada registro é DATAAGENDAMENTO, exceto na
        linha do tempo de concluídos, que usa DATACONCLUSAO (mais coerente
        com "quando foi concluído").
        """
        rows = self._excel_read_all()

        if ano and mes:
            meses_alvo = [(int(ano), int(mes))]
        elif ano:
            meses_alvo = [(int(ano), m) for m in range(1, 13)]
        else:
            meses_alvo = self._ultimos_n_meses(6)

        labels_meses = [f"{MESES_PT[m - 1]}/{str(a)[2:]}" for a, m in meses_alvo]

        def filtrar_por_mes(base_rows, campo_data, ano_alvo, mes_alvo):
            resultado = []
            for r in base_rows:
                d = _parse_date(r.get(campo_data))
                if not d:
                    continue
                ano_r, mes_r, _dia_r = d.split("-")
                if int(ano_r) == ano_alvo and int(mes_r) == mes_alvo:
                    resultado.append(r)
            return resultado

        # 1) Linha do tempo de concluídos (por DATACONCLUSAO)
        dados_concluidos = []
        for ano_m, mes_m in meses_alvo:
            subset = filtrar_por_mes(rows, "DATACONCLUSAO", ano_m, mes_m)
            dados_concluidos.append(
                sum(1 for r in subset if _strip_accents(r.get("STATUS")) == _strip_accents("Concluído"))
            )
        concluidos_timeline = {"labels": labels_meses, "data": dados_concluidos}

        # 2) Metragem por mês e tecnologia (por DATAAGENDAMENTO)
        tecnologias_presentes = sorted({
            (r.get("TECNOLOGIA") or "").strip() or "Não informado" for r in rows
        })
        metragem_datasets = []
        for tecnologia in tecnologias_presentes:
            valores = []
            for ano_m, mes_m in meses_alvo:
                subset = filtrar_por_mes(rows, "DATAAGENDAMENTO", ano_m, mes_m)
                total = sum(
                    _parse_metragem(r.get("METRAGEM"))
                    for r in subset
                    if ((r.get("TECNOLOGIA") or "").strip() or "Não informado") == tecnologia
                )
                valores.append(round(total, 1))
            metragem_datasets.append({"label": tecnologia, "data": valores})
        metragem_por_mes_tecnologia = {"labels": labels_meses, "datasets": metragem_datasets}

        # 2b) Total por Tipo de Cabo, por mês (por DATAAGENDAMENTO)
        tipos_cabo_presentes = sorted({
            (r.get("TIPOCABO") or "").strip() or "Não informado" for r in rows
        })
        tipocabo_datasets = []
        for tipo_cabo in tipos_cabo_presentes:
            valores = []
            for ano_m, mes_m in meses_alvo:
                subset = filtrar_por_mes(rows, "DATAAGENDAMENTO", ano_m, mes_m)
                total = sum(
                    1 for r in subset
                    if ((r.get("TIPOCABO") or "").strip() or "Não informado") == tipo_cabo
                )
                valores.append(total)
            tipocabo_datasets.append({"label": tipo_cabo, "data": valores})
        por_mes_tipocabo = {"labels": labels_meses, "datasets": tipocabo_datasets}

        # Junta todos os registros de todos os meses da janela, para as
        # agregações "totais no período" abaixo (executado por, cliente,
        # status, drafts).
        rows_na_janela = []
        for ano_m, mes_m in meses_alvo:
            rows_na_janela.extend(filtrar_por_mes(rows, "DATAAGENDAMENTO", ano_m, mes_m))

        def contar(campo, base_rows):
            contagem = {}
            for r in base_rows:
                chave = r.get(campo) or "Não informado"
                contagem[chave] = contagem.get(chave, 0) + 1
            return sorted(contagem.items(), key=lambda kv: kv[1], reverse=True)

        # 3) Total por Executado Por
        executado_ordenado = contar("EXECUTADOPOR", rows_na_janela)[:12]
        por_executado_por = {
            "labels": [k for k, _ in executado_ordenado],
            "data": [v for _, v in executado_ordenado],
        }

        # 4) Cliente com mais atividades
        cliente_ordenado = contar("CLIENTE", rows_na_janela)
        cliente_top_nome, cliente_top_total = cliente_ordenado[0] if cliente_ordenado else ("—", 0)

        # 5) Total por Status
        status_ordenado = contar("STATUS", rows_na_janela)
        por_status = {
            "labels": [k for k, _ in status_ordenado],
            "data": [v for _, v in status_ordenado],
        }

        # 6) Drafts criados por mês (registros com NUMDRAFT preenchido)
        dados_drafts = []
        for ano_m, mes_m in meses_alvo:
            subset = filtrar_por_mes(rows, "DATAAGENDAMENTO", ano_m, mes_m)
            dados_drafts.append(sum(1 for r in subset if (r.get("NUMDRAFT") or "").strip()))
        drafts_por_mes = {"labels": labels_meses, "data": dados_drafts}

        return {
            "meses_labels": labels_meses,
            "total_registros_periodo": len(rows_na_janela),
            "concluidos_timeline": concluidos_timeline,
            "metragem_por_mes_tecnologia": metragem_por_mes_tecnologia,
            "por_mes_tipocabo": por_mes_tipocabo,
            "por_executado_por": por_executado_por,
            "cliente_top": {"cliente": cliente_top_nome, "total": cliente_top_total},
            "por_status": por_status,
            "drafts_por_mes": drafts_por_mes,
        }

    def analytics_export(self, ano=None, mes=None, grafico="geral"):
        """Retorna as linhas (registros completos) que alimentam um gráfico
        específico da página de Análises, respeitando a mesma janela de
        tempo (ano/mês) usada em analytics(). Usado para o botão de baixar
        Excel em cada card de gráfico."""
        rows = self._excel_read_all()

        if ano and mes:
            meses_alvo = [(int(ano), int(mes))]
        elif ano:
            meses_alvo = [(int(ano), m) for m in range(1, 13)]
        else:
            meses_alvo = self._ultimos_n_meses(6)

        def filtrar_por_mes(base_rows, campo_data, ano_alvo, mes_alvo):
            resultado = []
            for r in base_rows:
                d = _parse_date(r.get(campo_data))
                if not d:
                    continue
                ano_r, mes_r, _dia_r = d.split("-")
                if int(ano_r) == ano_alvo and int(mes_r) == mes_alvo:
                    resultado.append(r)
            return resultado

        rows_na_janela = []
        for ano_m, mes_m in meses_alvo:
            rows_na_janela.extend(filtrar_por_mes(rows, "DATAAGENDAMENTO", ano_m, mes_m))

        if grafico == "concluidos_timeline":
            resultado = []
            for ano_m, mes_m in meses_alvo:
                subset = filtrar_por_mes(rows, "DATACONCLUSAO", ano_m, mes_m)
                resultado.extend(
                    r for r in subset if _strip_accents(r.get("STATUS")) == _strip_accents("Concluído")
                )
            return resultado

        if grafico == "drafts_por_mes":
            return [r for r in rows_na_janela if (r.get("NUMDRAFT") or "").strip()]

        # metragem_por_mes_tecnologia, por_mes_tipocabo, por_executado_por,
        # por_status (e qualquer outro/"geral") usam todos os registros da
        # janela, já que são apenas agrupamentos diferentes do mesmo total.
        return rows_na_janela

    def status_arquivo(self):
        """Diagnóstico: informa se o Excel foi localizado, onde, e quais
        abas ele tem (comparando com a aba configurada) — sem lançar
        exceção, útil para conferir rapidamente configuração/sincronização."""
        try:
            caminho = self._resolver_caminho_excel()
        except DataClientError as exc:
            return {"disponivel": False, "caminho": "", "mensagem": exc.message}

        abas = self._listar_abas(caminho)
        abas_lista = [a.strip() for a in abas.split(",")] if "," in abas or abas else [abas]
        aba_configurada_existe = self.excel_sheet in abas_lista

        return {
            "disponivel": True,
            "caminho": caminho,
            "aba_configurada": self.excel_sheet,
            "abas_no_arquivo": abas,
            "aba_configurada_existe": aba_configurada_existe,
            "mensagem": "Excel localizado." if aba_configurada_existe
                else f"Arquivo localizado, mas a aba '{self.excel_sheet}' não existe nele.",
        }

    def get_item(self, item_id):
        rows = self._excel_read_all()
        for row in rows:
            if str(row.get("IDCLIENTE")) == str(item_id):
                return row
        raise DataClientError(f"Registro com IDCLIENTE={item_id} não encontrado.", status_code=404)

    def create_item(self, data):
        data = self._sanitize_and_validate(data, is_new=True)
        rows = self._excel_read_all()
        if any(str(r.get("IDCLIENTE")) == str(data.get("IDCLIENTE")) for r in rows):
            raise DataClientError(f"Já existe um registro com IDCLIENTE={data.get('IDCLIENTE')}.", status_code=409)
        new_row = {f: data.get(f, "") for f in FIELDS}
        rows.append(new_row)
        self._excel_write_all(rows)
        logger.info("Registro criado: IDCLIENTE=%s", data.get("IDCLIENTE"))
        return new_row

    def update_item(self, item_id, data):
        data = self._sanitize_and_validate(data, is_new=False)
        rows = self._excel_read_all()
        found = False
        for row in rows:
            if str(row.get("IDCLIENTE")) == str(item_id):
                row.update(data)
                found = True
                break
        if not found:
            raise DataClientError(f"Registro com IDCLIENTE={item_id} não encontrado.", status_code=404)
        self._excel_write_all(rows)
        logger.info("Registro atualizado: IDCLIENTE=%s", item_id)
        return self.get_item(item_id)

    def delete_item(self, item_id):
        rows = self._excel_read_all()
        new_rows = [r for r in rows if str(r.get("IDCLIENTE")) != str(item_id)]
        if len(new_rows) == len(rows):
            raise DataClientError(f"Registro com IDCLIENTE={item_id} não encontrado.", status_code=404)
        self._excel_write_all(new_rows)
        logger.info("Registro excluído: IDCLIENTE=%s", item_id)
        return {"deleted": item_id}

    def dashboard_aggregates(self, filters=None):
        """Retorna agregações prontas para Chart.js + KPIs, respeitando filtros opcionais."""
        filters = filters or {}
        rows = self._excel_read_all()
        rows = self._apply_filters(rows, filters)

        def group_count(field, base_rows=None):
            base_rows = rows if base_rows is None else base_rows
            counts = {}
            for row in base_rows:
                key = row.get(field) or "Não informado"
                counts[key] = counts.get(key, 0) + 1
            sorted_items = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)[:12]
            return {
                "labels": [k for k, _ in sorted_items],
                "data": [v for _, v in sorted_items],
            }

        status_counts = group_count("STATUS")
        kpis = {
            "total": len(rows),
            "por_status": {k: v for k, v in zip(status_counts["labels"], status_counts["data"])},
        }

        concluidos = [r for r in rows if _strip_accents(r.get("STATUS")) == _strip_accents("Concluído")]

        if filters:
            # Há filtro(s) ativo(s): mostra os concluídos conforme o filtro aplicado
            # (rows já está filtrado por _apply_filters acima).
            concluidos_periodo = concluidos
            kpis["concluidos_label"] = "Concluídos (filtro)"
        else:
            # Sem filtro: por padrão, mostra apenas os concluídos do mês corrente.
            hoje = datetime.today()
            ano_atual, mes_atual = hoje.year, hoje.month

            def _no_mes_atual(row):
                d = _parse_date(row.get("DATACONCLUSAO"))
                if not d:
                    return False
                ano, mes, _dia = d.split("-")
                return int(ano) == ano_atual and int(mes) == mes_atual

            concluidos_periodo = [r for r in concluidos if _no_mes_atual(r)]
            kpis["concluidos_label"] = "Concluídos no mês"

        kpis["concluidos_total"] = len(concluidos_periodo)
        kpis["novos_total"] = sum(
            1 for r in rows if _strip_accents(r.get("STATUS")) == _strip_accents("Novo")
        )

        kpis["pendente_agendamento_total"] = sum(
            1 for r in rows if _strip_accents(r.get("STATUS")) == _strip_accents("PENDENTE AGENDAMENTO")
        )

        # Concluídos por mês, nos últimos 5 meses (incluindo o mês atual),
        # respeitando os filtros aplicados (mesma base "rows" dos demais
        # gráficos).
        concluidos_todos = [r for r in rows if _strip_accents(r.get("STATUS")) == _strip_accents("Concluído")]
        meses_alvo = self._ultimos_n_meses(5)
        labels_meses = [f"{MESES_PT[mes - 1]}/{str(ano)[2:]}" for ano, mes in meses_alvo]
        dados_meses = []
        for ano, mes in meses_alvo:
            total_mes = 0
            for r in concluidos_todos:
                d = _parse_date(r.get("DATACONCLUSAO"))
                if not d:
                    continue
                ano_r, mes_r, _dia_r = d.split("-")
                if int(ano_r) == ano and int(mes_r) == mes:
                    total_mes += 1
            dados_meses.append(total_mes)
        concluidos_5_meses = {"labels": labels_meses, "data": dados_meses}

        # "Por Executado Por": por padrão, mostra apenas o mês vigente
        # (usando DATAAGENDAMENTO); se houver filtro ativo, respeita o
        # filtro em vez do mês (mesma lógica já usada para "Concluídos").
        if filters:
            rows_executadopor = rows
        else:
            hoje = datetime.today()
            ano_atual, mes_atual = hoje.year, hoje.month

            def _no_mes_atual_agendamento(row):
                d = _parse_date(row.get("DATAAGENDAMENTO"))
                if not d:
                    return False
                ano, mes, _dia = d.split("-")
                return int(ano) == ano_atual and int(mes) == mes_atual

            rows_executadopor = [r for r in rows if _no_mes_atual_agendamento(r)]

        return {
            "kpis": kpis,
            "concluidos_5_meses": concluidos_5_meses,
            "por_executadopor": group_count("EXECUTADOPOR", rows_executadopor),
            "por_status": status_counts,
            "por_data_conclusao": self._group_by_date(concluidos_periodo, "DATACONCLUSAO"),
        }

    def items_by_date(self, date_str, date_field="DATAAGENDAMENTO", status=None):
        """Retorna todos os itens cujo campo de data indicado (por padrão
        DATAAGENDAMENTO, mas também aceita DATACONCLUSAO) seja igual a
        date_str (YYYY-MM-DD), opcionalmente filtrando por STATUS.
        Já projetados apenas com as colunas usadas no modal de detalhe."""
        if date_field not in DATE_FIELDS:
            raise DataClientError(f"Campo de data inválido: {date_field}")
        rows = self._excel_read_all()
        cols = ["IDCLIENTE", "CLIENTE", "ENDERECO", "CIDADE", "TECNOLOGIA", "VT",
                "DATAAGENDAMENTO", "DATACONCLUSAO", "STATUS", "TIPOCABO",
                "METRAGEM", "NUMDRAFT", "ROTA"]
        result = []
        for row in rows:
            if _parse_date(row.get(date_field)) != date_str:
                continue
            if status and _strip_accents(row.get("STATUS")) != _strip_accents(status):
                continue
            result.append({c: row.get(c) for c in cols})
        return result

    # ------------------------------------------------------------------
    # Filtros / ordenação
    # ------------------------------------------------------------------
    def _apply_filters(self, rows, filters):
        cliente = filters.get("cliente")
        idcliente = filters.get("id")
        cidade = filters.get("cidade")
        executadopor = filters.get("executadopor")
        status = filters.get("status")
        mes = filters.get("mes")
        data_inicio = filters.get("data_inicio")
        data_fim = filters.get("data_fim")
        busca = filters.get("q")

        def match(row):
            if cliente and _strip_accents(cliente) not in _strip_accents(row.get("CLIENTE")):
                return False
            if idcliente and str(idcliente) not in str(row.get("IDCLIENTE", "")):
                return False
            if cidade and _strip_accents(cidade) not in _strip_accents(row.get("CIDADE")):
                return False
            if executadopor and _strip_accents(executadopor) not in _strip_accents(row.get("EXECUTADOPOR")):
                return False
            if status and _strip_accents(status) != _strip_accents(row.get("STATUS")):
                return False
            if mes:
                if _strip_accents(row.get("STATUS")) != _strip_accents("Concluído"):
                    return False
                d = _parse_date(row.get("DATACONCLUSAO"))
                if not d:
                    return False
                _ano_d, mes_d, _dia_d = d.split("-")
                if int(mes_d) != int(mes):
                    return False
            if data_inicio or data_fim:
                d = _parse_date(row.get("DATAAGENDAMENTO"))
                if not d:
                    return False
                if data_inicio and d < data_inicio:
                    return False
                if data_fim and d > data_fim:
                    return False
            if busca:
                haystack = " ".join(str(row.get(f, "")) for f in FIELDS)
                if _strip_accents(busca) not in _strip_accents(haystack):
                    return False
            return True

        return [r for r in rows if match(r)]

    def _apply_sort(self, rows, sort):
        if not sort:
            return rows
        try:
            field, direction = sort.split(":")
        except ValueError:
            field, direction = sort, "asc"
        field = field.upper()
        if field not in FIELDS:
            return rows
        reverse = direction.lower() == "desc"

        def sort_key(row):
            value = row.get(field)
            if value is None:
                return ""
            return str(value).lower()

        return sorted(rows, key=sort_key, reverse=reverse)

    def _group_by_date(self, rows, field):
        counts = {}
        for row in rows:
            d = _parse_date(row.get(field))
            if not d:
                continue
            counts[d] = counts.get(d, 0) + 1
        ordered = sorted(counts.items(), key=lambda kv: kv[0])
        return {"labels": [k for k, _ in ordered], "data": [v for _, v in ordered]}

    @staticmethod
    def _ultimos_n_meses(n):
        """Retorna uma lista de tuplas (ano, mês) com os últimos n meses,
        incluindo o mês atual, do mais antigo para o mais recente."""
        hoje = datetime.today()
        ano, mes = hoje.year, hoje.month
        meses = []
        for _ in range(n):
            meses.append((ano, mes))
            mes -= 1
            if mes == 0:
                mes = 12
                ano -= 1
        return list(reversed(meses))

    def _sanitize_and_validate(self, data, is_new):
        """Sanitiza strings e valida formato de datas. Lança DataClientError se inválido."""
        clean = {}
        for field in FIELDS:
            if field not in data:
                continue
            value = data[field]
            if isinstance(value, str):
                value = value.strip()
            if field in DATE_FIELDS and value:
                try:
                    datetime.strptime(str(value)[:10], "%Y-%m-%d")
                except ValueError:
                    raise DataClientError(
                        f"Campo {field} deve estar no formato ISO yyyy-mm-dd. Valor recebido: {value}"
                    )
                value = str(value)[:10]
            clean[field] = value

        if is_new and not clean.get("IDCLIENTE"):
            raise DataClientError("IDCLIENTE é obrigatório para criar um registro.")
        if is_new and not clean.get("CLIENTE"):
            raise DataClientError("CLIENTE é obrigatório para criar um registro.")
        return clean

    # ------------------------------------------------------------------
    # Leitura / escrita do Excel
    # ------------------------------------------------------------------
    @staticmethod
    def _listar_abas(caminho):
        """Lista os nomes reais das abas de um arquivo Excel — usado para
        compor mensagens de erro mais úteis quando a aba configurada não
        existe."""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(caminho, read_only=True)
            try:
                return ", ".join(workbook.sheetnames) or "(nenhuma aba encontrada)"
            finally:
                workbook.close()
        except Exception:  # noqa: BLE001
            return "(não foi possível listar as abas)"

    def _excel_read_all(self):
        caminho = self._resolver_caminho_excel()
        with EXCEL_LOCK:
            try:
                df = pd.read_excel(caminho, sheet_name=self.excel_sheet, dtype=str)
            except PermissionError as exc:
                raise DataClientError(
                    "Não foi possível ler o Excel — feche o arquivo no Excel/OneDrive "
                    "e tente novamente.",
                    status_code=500,
                ) from exc
            except ValueError as exc:
                # Provavelmente a aba (sheet_name) configurada não existe neste
                # arquivo — lista as abas reais para facilitar a correção.
                abas_disponiveis = self._listar_abas(caminho)
                raise DataClientError(
                    f"A aba '{self.excel_sheet}' não foi encontrada no arquivo. "
                    f"Abas disponíveis: {abas_disponiveis}. Ajuste EXCEL_SHEET_NAME "
                    "no .env para o nome exato de uma dessas abas.",
                    status_code=500,
                ) from exc
            except Exception as exc:  # noqa: BLE001 - queremos capturar qualquer erro de leitura
                logger.exception("Erro lendo Excel")
                raise DataClientError(f"Erro ao ler o arquivo Excel: {exc}", status_code=500) from exc

        for field in FIELDS:
            if field not in df.columns:
                df[field] = ""
        df = df.fillna("")
        rows = df.to_dict(orient="records")

        # Normaliza datas para ISO (yyyy-mm-dd), aceitando tanto o formato
        # ISO quanto dd/mm/aaaa (comum quando editado manualmente no Excel
        # em pt-BR). Isso garante que filtros, ordenação e exibição no
        # frontend funcionem de forma consistente, seja qual for o formato
        # em que a data foi digitada na planilha.
        for row in rows:
            for field in DATE_FIELDS:
                normalized = _parse_date(row.get(field))
                row[field] = normalized or ""

        return rows

    def _excel_write_all(self, rows):
        caminho = self._resolver_caminho_excel()
        caminho_path = Path(caminho)
        df = pd.DataFrame(rows, columns=FIELDS)
        arquivo_temporario = None

        with EXCEL_LOCK:
            try:
                # Escreve primeiro num arquivo temporário na mesma pasta e só
                # substitui o arquivo real no final (os.replace é atômico).
                # Isso evita deixar o Excel corrompido/pela metade caso algo
                # falhe no meio da escrita (ex.: sincronização do OneDrive).
                with tempfile.NamedTemporaryFile(
                    prefix="redeb2b_tmp_", suffix=".xlsx",
                    dir=str(caminho_path.parent), delete=False,
                ) as tmp:
                    arquivo_temporario = tmp.name

                with pd.ExcelWriter(arquivo_temporario, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=self.excel_sheet, index=False)

                os.replace(arquivo_temporario, caminho)
            except PermissionError as exc:
                raise DataClientError(
                    "Não foi possível salvar o Excel — feche o arquivo no Excel, "
                    "confirme a sincronização do OneDrive e tente novamente.",
                    status_code=500,
                ) from exc
            except Exception as exc:  # noqa: BLE001
                logger.exception("Erro escrevendo Excel")
                raise DataClientError(f"Erro ao gravar o arquivo Excel: {exc}", status_code=500) from exc
            finally:
                if arquivo_temporario and os.path.exists(arquivo_temporario):
                    try:
                        os.remove(arquivo_temporario)
                    except OSError:
                        pass
