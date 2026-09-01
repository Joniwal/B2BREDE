"""CRUD conservador: alterar células/tabela existentes, nunca recriar uma aba.

Perfil: XLSX comum, 21 campos de entrada, chave única e estável. Fórmulas fora
dos campos editados são mantidas; o openpyxl NÃO calcula fórmulas/pivôs.
Objetos não homologados são recusados antes de qualquer commit.
"""
from __future__ import annotations

import json
import re
import warnings
import zipfile
from copy import copy, deepcopy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator, TranslatorError
from openpyxl.utils.cell import range_boundaries, get_column_letter

from excel_io import ExcelSafetyError, digest, transact


def text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def inventory(content: bytes) -> dict:
    """Triagem, não certificação de fidelidade. Nenhum arquivo é salvo aqui."""
    risky = []
    with zipfile.ZipFile(BytesIO(content)) as archive:
        names = archive.namelist()
        for name in names:
            lower = name.lower()
            if any(part in lower for part in (
                "vbaproject", "slicer", "pivot", "activex/", "ctrlprops/",
                "connections.xml", "querytables/", "customxml/", "embeddings/",
                "threadedcomments/", "persons/", "externallinks/", "xl/model/",
                "_xmlsignatures/", "customui/", "vml", "macrosheets/", "xl/metadata.xml",
            )):
                risky.append(name)
            if name.endswith(".xml") and name.startswith("xl/"):
                root = ET.fromstring(archive.read(name))
                tags = {node.tag.rsplit("}", 1)[-1] for node in root.iter()}
                if "extLst" in tags or (name.startswith("xl/drawings/") and tags & {"sp", "grpSp", "cxnSp", "contentPart"}):
                    risky.append(name + " (extensões/formas)")
        return {"parts": sorted(names), "risks": sorted(set(risky))}


def guard_roundtrip(original, staged):
    """Impede sumiço de partes OOXML. Não substitui teste visual no Excel."""
    before = set(inventory(original)["parts"])
    with zipfile.ZipFile(staged) as archive:
        after = set(archive.namelist())
    # Essas duas partes podem ser normalizadas/recriadas legitimamente.
    allowed_removed = {"xl/sharedStrings.xml", "xl/calcChain.xml"}
    lost = before - after - allowed_removed
    if lost:
        raise ExcelSafetyError("Partes do Excel seriam perdidas: " + ", ".join(sorted(lost)), "EXCEL_OBJECT_LOSS")


class SafeExcel:
    def __init__(self, path, sheet, fields, *, key="_ITEM_ID", table=None,
                 header_row=1, date_fields=(), backup_dir=None,
                 lock_timeout=15, attempts=4, read_only=False):
        self.path = Path(path).resolve()
        self.sheet = sheet
        self.fields = tuple(fields)
        self.key = key.upper()
        self.table = table
        self.header_row = int(header_row)
        self.dates = set(date_fields)
        self.backup_dir = backup_dir
        self.lock_timeout, self.attempts = lock_timeout, attempts
        self.read_only = read_only

    def _layout(self, workbook):
        if self.sheet not in workbook.sheetnames:
            raise ExcelSafetyError(f"Aba '{self.sheet}' não existe. Nenhuma aba foi criada.", "EXCEL_SCHEMA")
        ws = workbook[self.sheet]
        header, start_col, end_col, end_row = self.header_row, 1, ws.max_column, ws.max_row
        table = None
        if self.table:
            if self.table not in ws.tables:
                raise ExcelSafetyError(f"Tabela '{self.table}' não encontrada.", "EXCEL_SCHEMA")
            table = ws.tables[self.table]
        elif len(ws.tables) == 1:
            table = next(iter(ws.tables.values()))
        elif len(ws.tables) > 1:
            raise ExcelSafetyError("Há várias tabelas; configure EXCEL_TABLE.", "EXCEL_SCHEMA")
        if table is not None:
            start_col, header, end_col, end_row = range_boundaries(table.ref)
            if table.headerRowCount == 0:
                raise ExcelSafetyError("A tabela precisa de cabeçalhos.", "EXCEL_SCHEMA")
            end_row -= int(table.totalsRowCount or bool(table.totalsRowShown))
        headers = {}
        for col in range(start_col, end_col + 1):
            name = text(ws.cell(header, col).value).upper()
            if name:
                if name in headers:
                    raise ExcelSafetyError(f"Cabeçalho duplicado: {name}", "EXCEL_SCHEMA")
                headers[name] = col
        if self.key not in headers:
            raise ExcelSafetyError(f"Chave '{self.key}' ausente. Prepare-a numa janela de manutenção; leitura não altera o arquivo.", "EXCEL_SCHEMA")
        return ws, headers, header, end_row, table

    def _rows(self, ws, headers, header, end_row):
        rows = {}
        for row in range(header + 1, end_row + 1):
            key_cell = ws.cell(row, headers[self.key])
            item_id = text(key_cell.value)
            if key_cell.data_type == "f":
                raise ExcelSafetyError("A chave não pode ser uma fórmula.", "EXCEL_SCHEMA")
            if not item_id:
                # Linhas reservadas ou apagadas não são registros. Dados órfãos
                # não são corrigidos silenciosamente nem recebem IDs por posição.
                if any(ws.cell(row, headers[f]).value not in (None, "")
                       for f in self.fields if f in headers):
                    raise ExcelSafetyError(f"Linha {row} tem dados sem chave.", "EXCEL_SCHEMA")
                continue
            if item_id in rows:
                raise ExcelSafetyError(f"Chave duplicada na linha {row}.", "EXCEL_SCHEMA")
            rows[item_id] = row
        return rows

    def _item(self, ws, headers, row):
        result = {f: text(ws.cell(row, headers[f]).value) if f in headers else "" for f in self.fields}
        result["id"] = text(ws.cell(row, headers[self.key]).value)
        serialized = json.dumps(result, sort_keys=True, ensure_ascii=False).encode("utf-8")
        result["_etag"] = digest(serialized)
        return result

    def read_rows(self):
        """Snapshot em memória; NUNCA chama save(), migração, mkdir ou lock."""
        try:
            content = self.path.read_bytes()
        except FileNotFoundError as exc:
            raise ExcelSafetyError("Arquivo Excel não encontrado.", "EXCEL_NOT_FOUND", 404) from exc
        # data_only=False impede confundir fórmulas com valores persistíveis.
        wb = load_workbook(BytesIO(content), data_only=False, keep_links=True, rich_text=True)
        cached = None
        try:
            ws, headers, header, end_row, _ = self._layout(wb)
            rows = self._rows(ws, headers, header, end_row)
            items = []
            for row in rows.values():
                item = self._item(ws, headers, row)
                for field in self.fields:
                    if field in headers and ws.cell(row, headers[field]).data_type == "f":
                        if cached is None:
                            cached = load_workbook(BytesIO(content), data_only=True, read_only=True)
                        value = cached[self.sheet].cell(row, headers[field]).value
                        if value is None:
                            raise ExcelSafetyError("Fórmula sem resultado armazenado. Recalcule no Excel antes de consultar.", "EXCEL_RECALC_REQUIRED")
                        item[field] = text(value)
                items.append(item)
            return items
        finally:
            wb.close()
            if cached is not None:
                cached.close()

    def get(self, item_id):
        for item in self.read_rows():
            if item["id"] == str(item_id):
                return item
        raise ExcelSafetyError("Registro não encontrado.", "NOT_FOUND", 404)

    def _convert(self, field, value):
        if value is None or value == "":
            return None
        if isinstance(value, (list, dict, tuple, set)):
            raise ExcelSafetyError(f"Valor inválido em {field}.", "VALIDATION_ERROR", 422)
        if field in self.dates:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(value)):
                raise ExcelSafetyError(f"{field}: utilize AAAA-MM-DD.", "VALIDATION_ERROR", 422)
            try:
                return date.fromisoformat(value)
            except ValueError as exc:
                raise ExcelSafetyError(f"Data inválida em {field}.", "VALIDATION_ERROR", 422) from exc
        result = str(value).strip()
        if len(result) > 4000 or re.search(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", result):
            raise ExcelSafetyError(f"Texto inválido em {field}.", "VALIDATION_ERROR", 422)
        return result

    def _assign(self, ws, row, col, field, value):
        cell = ws.cell(row, col)
        if cell.data_type == "f":
            raise ExcelSafetyError(f"{cell.coordinate} contém fórmula; não pode ser sobrescrita.", "EXCEL_FORMULA_PROTECTED")
        # Atualizar .value não redefine fonte/fill/borda/alinhamento/proteção.
        # Datas podem provocar formatação automática da biblioteca: restauramos
        # o formato existente, salvo General, onde ISO torna a nova data legível.
        number_format = cell.number_format
        cell.value = value
        if isinstance(value, str):
            cell.data_type = "s"  # '=...' é texto literal, não fórmula injetada.
        cell.number_format = "yyyy-mm-dd" if isinstance(value, date) and number_format == "General" else number_format

    def _extend_table(self, ws, headers, header, end_row, table):
        """Amplia uma linha da mesma tabela sem recriar tabela, aba ou workbook.

        O estilo estruturado continua pertencendo ao mesmo objeto Table. Estilos
        de célula/altura e fórmulas de colunas auxiliares são copiados da última
        linha usando referências traduzidas. Conteúdo logo abaixo bloqueia a
        operação para que nunca seja sobrescrito silenciosamente.
        """
        left, top, right, bottom = range_boundaries(table.ref)
        totals = int(table.totalsRowCount or bool(table.totalsRowShown))
        if totals:
            raise ExcelSafetyError(
                "Tabela com linha de totais: amplie-a pelo Excel nativo ou Office Script homologado.",
                "EXCEL_TABLE_TOTALS",
            )
        if top != header or bottom != end_row:
            raise ExcelSafetyError("Intervalo da tabela é inconsistente.", "EXCEL_SCHEMA")

        target = bottom + 1
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= target <= merged.max_row and not (
                merged.max_col < left or merged.min_col > right
            ):
                raise ExcelSafetyError(
                    "A linha abaixo da tabela contém células mescladas; nada foi alterado.",
                    "EXCEL_TABLE_BLOCKED",
                )
        for other in ws.tables.values():
            if other.name == table.name:
                continue
            other_left, other_top, other_right, other_bottom = range_boundaries(other.ref)
            if other_top <= target <= other_bottom and not (
                other_right < left or other_left > right
            ):
                raise ExcelSafetyError(
                    "Outra tabela ocupa a próxima linha; nada foi alterado.",
                    "EXCEL_TABLE_BLOCKED",
                )
        if any(ws.cell(target, col).value not in (None, "") for col in range(left, right + 1)):
            raise ExcelSafetyError(
                "Há conteúdo logo abaixo da tabela; nada foi sobrescrito.",
                "EXCEL_TABLE_BLOCKED",
            )

        old_ref = table.ref
        new_ref = f"{get_column_letter(left)}{top}:{get_column_letter(right)}{target}"
        if table.autoFilter is not None and table.autoFilter.ref not in (None, old_ref):
            raise ExcelSafetyError(
                "O filtro interno da tabela tem um intervalo independente; nada foi alterado.",
                "EXCEL_TABLE_BLOCKED",
            )

        source = end_row if end_row > header else None
        if source is not None:
            names_by_column = {column: name for name, column in headers.items()}
            managed = {*self.fields, self.key}
            for col in range(left, right + 1):
                source_cell = ws.cell(source, col)
                target_cell = ws.cell(target, col)
                # Uma célula vazia já pré-formatada fora da tabela é preservada.
                if not target_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.data_type == "f" and names_by_column.get(col) not in managed:
                    try:
                        target_cell.value = Translator(
                            source_cell.value, origin=source_cell.coordinate
                        ).translate_formula(target_cell.coordinate)
                    except TranslatorError as exc:
                        raise ExcelSafetyError(
                            f"Não foi possível estender a fórmula auxiliar de {source_cell.coordinate}.",
                            "EXCEL_FORMULA_TRANSLATION",
                        ) from exc
            if ws.row_dimensions[target].height is None:
                ws.row_dimensions[target].height = ws.row_dimensions[source].height

            # Mantém dropdowns/validações e formatação condicional que atingiam
            # a última linha-modelo. As regras existentes não são substituídas.
            for validation in ws.data_validations.dataValidation:
                for col in range(left, right + 1):
                    source_coordinate = ws.cell(source, col).coordinate
                    if source_coordinate in validation.sqref:
                        validation.add(ws.cell(target, col).coordinate)

            conditional_ranges = [
                (conditional, list(ws.conditional_formatting._cf_rules[conditional]))
                for conditional in ws.conditional_formatting
            ]
            for conditional, rules in conditional_ranges:
                for col in range(left, right + 1):
                    source_coordinate = ws.cell(source, col).coordinate
                    if source_coordinate not in conditional.sqref:
                        continue
                    target_coordinate = ws.cell(target, col).coordinate
                    for rule in rules:
                        copied_rule = deepcopy(rule)
                        copied_rule.formula = [
                            Translator(formula, origin=source_coordinate).translate_formula(target_coordinate)
                            if isinstance(formula, str) and formula.startswith("=")
                            else formula
                            for formula in (copied_rule.formula or [])
                        ]
                        ws.conditional_formatting.add(target_coordinate, copied_rule)

        table.ref = new_ref
        if table.autoFilter is not None:
            table.autoFilter.ref = new_ref
        if ws.auto_filter.ref == old_ref:
            ws.auto_filter.ref = new_ref
        return target

    def _run(self, operation):
        if self.read_only:
            raise ExcelSafetyError("Modo somente leitura. Encaminhe alterações ao escritor único.", "EXCEL_READ_ONLY", 403)
        if self.path.suffix.lower() != ".xlsx":
            raise ExcelSafetyError("Este adaptador grava apenas XLSX homologado; use Excel nativo para XLSM/XLSB.", "EXCEL_COMPLEX_FILE")

        def transform(original, staged):
            report = inventory(original)
            if report["risks"]:
                raise ExcelSafetyError("Objetos complexos detectados. Gravação bloqueada: " + ", ".join(report["risks"][:6]), "EXCEL_COMPLEX_FILE")
            # Avisos de recursos removidos nunca são ignorados.
            with warnings.catch_warnings():
                warnings.simplefilter("error", UserWarning)
                wb = load_workbook(BytesIO(original), data_only=False, keep_links=True, rich_text=True)
                try:
                    ws, headers, header, end_row, table = self._layout(wb)
                    if ws.protection.sheet:
                        raise ExcelSafetyError("Aba protegida; homologação necessária.", "EXCEL_PROTECTED")
                    managed_cols = [headers[f] for f in (*self.fields, self.key) if f in headers]
                    for merged in ws.merged_cells.ranges:
                        if merged.max_row > header and merged.min_row <= end_row and any(merged.min_col <= col <= merged.max_col for col in managed_cols):
                            raise ExcelSafetyError("Há células mescladas na área de dados.", "EXCEL_SCHEMA")
                    rows = self._rows(ws, headers, header, end_row)
                    result, changed = operation(ws, headers, header, end_row, table, rows)
                    if changed:
                        wb.save(staged)
                        guard_roundtrip(original, staged)
                        # Reabre o resultado antes do commit para detectar ZIP/XML inválido.
                        check = load_workbook(staged, read_only=True, data_only=False)
                        check.close()
                    return result, changed
                finally:
                    wb.close()
        try:
            return transact(self.path, transform, backup_dir=self.backup_dir,
                            lock_timeout=self.lock_timeout, attempts=self.attempts)
        except UserWarning as exc:
            raise ExcelSafetyError("Biblioteca sinalizou perda de recurso; original mantido.", "EXCEL_OBJECT_LOSS") from exc

    def update(self, item_id, data, *, expected_version=None):
        unknown = set(data) - set(self.fields)
        if unknown:
            raise ExcelSafetyError("Campos desconhecidos: " + ", ".join(sorted(unknown)), "VALIDATION_ERROR", 422)
        values = {f: self._convert(f, v) for f, v in data.items()}

        def edit(ws, headers, header, end_row, table, rows):
            if str(item_id) not in rows:
                raise ExcelSafetyError("Registro não encontrado.", "NOT_FOUND", 404)
            row = rows[str(item_id)]
            current = self._item(ws, headers, row)
            if expected_version and expected_version.strip('"') != current["_etag"]:
                raise ExcelSafetyError("Registro alterado por outro usuário. Reabra o modal.", "EXCEL_STALE_RECORD", 412)
            changed = False
            for field, value in values.items():
                if field not in headers:
                    raise ExcelSafetyError(f"Coluna {field} ausente; esquema não será alterado.", "EXCEL_SCHEMA")
                if field == self.key and text(value) != str(item_id):
                    raise ExcelSafetyError("A chave técnica é imutável.", "EXCEL_KEY_IMMUTABLE")
                cell = ws.cell(row, headers[field])
                if text(cell.value) != text(value):
                    self._assign(ws, row, headers[field], field, value)
                    changed = True
            return self._item(ws, headers, row), changed
        return self._run(edit)

    def create(self, data):
        unknown = set(data) - set(self.fields)
        if unknown:
            raise ExcelSafetyError("Campos desconhecidos.", "VALIDATION_ERROR", 422)
        values = {f: self._convert(f, data.get(f)) for f in self.fields}
        # UUID gerado uma vez por transação: retries de PermissionError não o mudam.
        item_id = text(values.get(self.key)) if self.key in self.fields else str(uuid4())
        if not item_id or not values.get("IDCLIENTE") or not values.get("CLIENTE"):
            raise ExcelSafetyError("IDCLIENTE, CLIENTE e chave são obrigatórios.", "VALIDATION_ERROR", 422)

        def add(ws, headers, header, end_row, table, rows):
            if item_id in rows:
                raise ExcelSafetyError("Chave já cadastrada.", "EXCEL_DUPLICATE_KEY")
            for f, value in values.items():
                if f not in headers and value is not None:
                    raise ExcelSafetyError(f"Coluna {f} ausente.", "EXCEL_SCHEMA")
            occupied = set(rows.values())
            row = next((r for r in range(header + 1, end_row + 1) if r not in occupied), end_row + 1)
            if row > end_row and table is not None:
                # Altera somente o ref da MESMA tabela. Nome, id, estilo,
                # colunas, filtros e objetos do workbook não são recriados.
                row = self._extend_table(ws, headers, header, end_row, table)
            if row > end_row:
                # Sem tabela: só expande uma aba tabular simples, sem ranges que
                # precisariam ser estendidos/reinterpretados automaticamente.
                if table is None and (ws.conditional_formatting or ws.data_validations.count or ws._charts or ws._pivots or ws.merged_cells.ranges):
                    raise ExcelSafetyError("Ampliação exige homologar ranges; reserve linhas de dados.", "EXCEL_TABLE_CAPACITY")
                if table is None and ws.auto_filter.ref:
                    left, top, right, bottom = range_boundaries(ws.auto_filter.ref)
                    if top != header or bottom != end_row or min(headers.values()) < left or max(headers.values()) > right:
                        raise ExcelSafetyError("AutoFilter possui intervalo independente; reserve linhas.", "EXCEL_TABLE_CAPACITY")
                    ws.auto_filter.ref = f"{get_column_letter(left)}{top}:{get_column_letter(right)}{row}"
                source_row = next(iter(rows.values()), None)
                if table is None and source_row:
                    # Fórmulas adicionais devem ter extensão deliberada no template.
                    if any(ws.cell(source_row, c).data_type == "f" for c in range(1, ws.max_column + 1)):
                        raise ExcelSafetyError("Reserve linhas com fórmulas antes de criar itens.", "EXCEL_TABLE_CAPACITY")
                    for col in headers.values():
                        ws.cell(row, col)._style = copy(ws.cell(source_row, col)._style)
                    ws.row_dimensions[row].height = ws.row_dimensions[source_row].height
            for field, value in values.items():
                if field in headers:
                    self._assign(ws, row, headers[field], field, value)
            self._assign(ws, row, headers[self.key], self.key, item_id)
            return self._item(ws, headers, row), True
        return self._run(add)

    def delete(self, item_id, *, expected_version=None):
        def remove(ws, headers, header, end_row, table, rows):
            if str(item_id) not in rows:
                raise ExcelSafetyError("Registro não encontrado.", "NOT_FOUND", 404)
            row = rows[str(item_id)]
            current = self._item(ws, headers, row)
            if expected_version and expected_version.strip('"') != current["_etag"]:
                raise ExcelSafetyError("Registro alterado por outro usuário. Recarregue.", "EXCEL_STALE_RECORD", 412)
            # Não delete_rows(): posições, fórmulas externas e objetos não se movem.
            # Colunas extras não gerenciadas (ex.: fórmulas de relatório) permanecem.
            for field in dict.fromkeys((*self.fields, self.key)):
                if field in headers:
                    self._assign(ws, row, headers[field], field, None)
            return None, True
        return self._run(remove)
