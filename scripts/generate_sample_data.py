"""Regenera a planilha de demonstração usada no modo Excel."""

from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

from sharepoint_client import FIELDS  # noqa: E402


def build_rows() -> list[dict[str, str]]:
    clients = ["Alfa Telecom", "Banco Horizonte", "Clínica Aurora", "Delta Logística", "Editora Central", "Fábrica Norte"]
    cities = ["São Paulo", "Campinas", "Santos", "Sorocaba", "Jundiaí", "São José dos Campos"]
    statuses = ["Pendente", "Agendado", "Em execução", "Concluído", "Concluído", "Cancelado"]
    executors = ["Equipe Ômega", "Equipe Leste", "Parceiro Sul", "Equipe Ômega", "Equipe Norte", "Parceiro Sul"]
    technologies = ["GPON", "Metro Ethernet", "FTTH", "IP MPLS", "GPON", "DWDM"]
    start = date.today() - timedelta(days=12)
    rows: list[dict[str, str]] = []
    for index in range(30):
        status = statuses[index % len(statuses)]
        scheduled = start + timedelta(days=index % 16)
        client = clients[index % len(clients)]
        row = {field: "" for field in FIELDS}
        row.update(
            IDCLIENTE=f"B2B-{1001 + index}",
            CLIENTE=client,
            ENDERECO=f"Av. Operacional, {120 + index}",
            CIDADE=cities[index % len(cities)],
            PRODUTO="Link dedicado",
            ATIVIDADE="Implantação" if index % 3 else "Ampliação",
            TECNOLOGIA=technologies[index % len(technologies)],
            VT=f"VT-{24000 + index}",
            DATADISPARO=(scheduled - timedelta(days=5)).isoformat(),
            RETORNOPCC=(scheduled - timedelta(days=3)).isoformat(),
            DATAAGENDAMENTO=scheduled.isoformat(),
            DATACONCLUSAO=scheduled.isoformat() if "Concluído" in status else "",
            OBSERVACAO="Janela confirmada com o cliente." if index % 2 else "Aguardando liberação de acesso.",
            STATUS=status,
            EXECUTADOPOR=executors[index % len(executors)],
            TIPOCABO="Fibra óptica 12FO" if index % 2 else "Fibra óptica 24FO",
            METRAGEM=str(80 + (index * 17) % 420),
            OBSERVACAOCONCLUSAO="Testes de potência e conectividade aprovados." if "Concluído" in status else "",
            NUMDRAFT=f"DR-{7000 + index}",
            ROTA=f"RT-{(index % 8) + 1:02d}",
            USUARIO="demo.operacao",
        )
        rows.append({"_ITEM_ID": str(index + 1), **row})
    return rows


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Cria uma NOVA demonstração; nunca sobrescreve o arquivo real.")
    parser.add_argument("--output", default=str(PROJECT_DIR / "data" / "REDE_B2B_NOVO_EXEMPLO.xlsx"))
    args = parser.parse_args()
    output = Path(args.output).resolve()
    if output.exists():
        parser.error("Destino já existe. Informe outro --output; nenhum dado foi substituído.")
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    columns = ["_ITEM_ID", *FIELDS]
    ws.append(columns)
    for item in build_rows():
        ws.append([item.get(c, "") for c in columns])
    for row in ws.iter_rows(min_row=2, max_row=101, max_col=len(columns)):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="5B2A86")
        cell.font = Font(color="FFFFFF", bold=True)
    table = Table(displayName="TabelaRede", ref="A1:V101")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "B2"
    # A chave fica visível para facilitar o conector Excel Online/OData.
    with output.open("xb") as destination:
        wb.save(destination)
    wb.close()
    print(f"Planilha criada: {output}")
