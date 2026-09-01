"""Inventário somente leitura; não exige abrir o Excel nem envia dados à nuvem."""
import argparse
import json
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from excel_safe import inventory
from excel_io import digest


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file")
    args = parser.parse_args()
    path = Path(args.file)
    content = path.read_bytes()
    report = inventory(content)
    wb = load_workbook(BytesIO(content), keep_links=True, data_only=False)
    try:
        report.update(sha256=digest(content), bytes=len(content), sheets=[{
            "name": ws.title, "rows": ws.max_row, "columns": ws.max_column,
            "tables": [{"name": t.name, "id": t.id, "range": t.ref} for t in ws.tables.values()],
            "charts": len(ws._charts), "pivots": len(ws._pivots),
            "merged": [str(x) for x in ws.merged_cells.ranges],
        } for ws in wb])
        print(json.dumps(report, ensure_ascii=True, indent=2))
    finally:
        wb.close()


if __name__ == "__main__":
    main()
