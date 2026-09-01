from __future__ import annotations

import multiprocessing
import zipfile
from concurrent.futures import ProcessPoolExecutor
from copy import copy

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

from app import create_app
from excel_io import ExcelSafetyError, digest, local_lock
from excel_safe import SafeExcel
from sharepoint_client import ExcelDataStore, FIELDS, DATE_FIELDS


def make_fixture(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    cols = ["_ITEM_ID", *FIELDS, "CALCULADO"]
    ws.append(cols)
    for number in (1, 2):
        item = {"_ITEM_ID": str(number), "IDCLIENTE": f"B2B-{number}", "CLIENTE": "Cliente " + str(number),
                "STATUS": "Agendado", "CIDADE": "Campinas", "DATAAGENDAMENTO": "2026-08-28"}
        ws.append([item.get(c) for c in cols])
    for row in range(2, 7):
        for col in range(1, len(cols) + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", size=10, color="552288")
            cell.fill = PatternFill("solid", fgColor="ECE2F4")
            cell.border = Border(bottom=Side(style="thin", color="DEDEDE"))
            cell.alignment = Alignment(vertical="center")
        ws.cell(row, 23, f'=IF(A{row}="",0,1)')
    ws.column_dimensions["C"].width = 32
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "B2"
    table = Table(displayName="TabelaRede", ref="A1:W6")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)
    validation = DataValidation(type="list", formula1='"Agendado,Concluído,Pendente"')
    ws.add_data_validation(validation)
    validation.add("O2:O6")
    ws.conditional_formatting.add("R2:R6", CellIsRule(operator="greaterThan", formula=["100"], fill=PatternFill("solid", fgColor="FFDDCC")))
    summary = wb.create_sheet("Resumo")
    summary.append(["Exemplo", "Total"])
    summary.append(["A", 1])
    summary.append(["B", 2])
    summary["D2"] = '=COUNTA(Dados!A2:A6)'
    summary["D2"].number_format = "0"
    chart = BarChart()
    chart.add_data(Reference(summary, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    summary.add_chart(chart, "F2")
    wb.defined_names.add(DefinedName("FaixaRede", attr_text="'Dados'!$A$1:$W$6"))
    wb.save(path)
    wb.close()


@pytest.fixture
def fixture_path(tmp_path):
    path = tmp_path / "REDE_B2B.xlsx"
    make_fixture(path)
    return path


def store(path, **kwargs):
    return SafeExcel(path, "Dados", FIELDS, table="TabelaRede", date_fields=DATE_FIELDS, **kwargs)


def structural_snapshot(path):
    wb = load_workbook(path, data_only=False)
    try:
        return {
            "sheets": wb.sheetnames,
            "styles": [(ws.title, c.coordinate, copy(c._style)) for ws in wb for row in ws for c in row],
            "formulas": [(ws.title, c.coordinate, c.value) for ws in wb for row in ws for c in row if c.data_type == "f"],
            "tables": [(ws.title, t.name, t.id, t.ref, str(t.tableStyleInfo)) for ws in wb for t in ws.tables.values()],
            "validation": str(wb["Dados"].data_validations),
            "cf": [(str(key), str(rules)) for key, rules in wb["Dados"].conditional_formatting._cf_rules.items()],
            "freeze": wb["Dados"].freeze_panes,
            "width": wb["Dados"].column_dimensions["C"].width,
            "height": wb["Dados"].row_dimensions[2].height,
            "names": str(wb.defined_names),
        }
    finally:
        wb.close()


def test_get_routes_never_change_file(fixture_path, tmp_path):
    repository = ExcelDataStore(fixture_path, table="TabelaRede")
    client = create_app({"TESTING": True, "LOG_DIR": str(tmp_path / "logs")}, repository).test_client()
    before, modified = digest(fixture_path.read_bytes()), fixture_path.stat().st_mtime_ns
    for _ in range(3):
        for route in ("/", "/api/items", "/api/items/1", "/api/dashboard", "/api/meta", "/api/health"):
            assert client.get(route).status_code == 200
    assert digest(fixture_path.read_bytes()) == before
    assert fixture_path.stat().st_mtime_ns == modified
    assert not fixture_path.with_name(fixture_path.name + ".lock").exists()
    assert not (fixture_path.parent / "_rede_backups").exists()


def test_patch_preserves_styles_formulas_tables_chart_and_other_sheets(fixture_path):
    before = structural_snapshot(fixture_path)
    with zipfile.ZipFile(fixture_path) as archive:
        chart = archive.read("xl/charts/chart1.xml")
    saved_bytes = fixture_path.read_bytes()
    result = store(fixture_path).update("1", {"STATUS": "Concluído"})
    assert result["STATUS"] == "Concluído"
    assert structural_snapshot(fixture_path) == before
    with zipfile.ZipFile(fixture_path) as archive:
        assert archive.read("xl/charts/chart1.xml") == chart
    backups = list((fixture_path.parent / "_rede_backups").glob("*.xlsx"))
    assert len(backups) == 1 and backups[0].read_bytes() == saved_bytes


def test_create_and_delete_preserve_fixed_table_and_formatted_slots(fixture_path):
    before = structural_snapshot(fixture_path)
    adapter = store(fixture_path)
    created = adapter.create({"IDCLIENTE": "NOVA", "CLIENTE": "Cliente novo"})
    assert len(adapter.read_rows()) == 3
    assert structural_snapshot(fixture_path) == before
    adapter.delete(created["id"])
    assert len(adapter.read_rows()) == 2
    assert structural_snapshot(fixture_path) == before
    assert adapter.get("2")["CLIENTE"] == "Cliente 2"


def test_formula_injection_remains_text(fixture_path):
    store(fixture_path).update("1", {"OBSERVACAO": '=HYPERLINK("https://example.invalid")'})
    wb = load_workbook(fixture_path)
    assert wb["Dados"]["N2"].data_type == "s"
    wb.close()


def test_existing_formula_cannot_be_overwritten(fixture_path):
    wb = load_workbook(fixture_path)
    wb["Dados"]["O2"] = '="Agendado"'
    wb.save(fixture_path)
    wb.close()
    original = fixture_path.read_bytes()
    with pytest.raises(ExcelSafetyError, match="fórmula"):
        store(fixture_path).update("1", {"STATUS": "Concluído"})
    assert fixture_path.read_bytes() == original


def test_duplicate_key_is_read_error_not_migration(fixture_path):
    wb = load_workbook(fixture_path)
    wb["Dados"]["A3"] = "1"
    wb.save(fixture_path)
    wb.close()
    original = fixture_path.read_bytes()
    with pytest.raises(ExcelSafetyError, match="duplicada"):
        store(fixture_path).read_rows()
    assert fixture_path.read_bytes() == original


def test_missing_file_never_auto_created(tmp_path):
    missing = tmp_path / "missing.xlsx"
    with pytest.raises(ExcelSafetyError):
        store(missing).read_rows()
    assert not missing.exists()


def test_noop_does_not_save(fixture_path):
    original = fixture_path.read_bytes()
    store(fixture_path).update("1", {"STATUS": "Agendado"})
    assert fixture_path.read_bytes() == original
    assert not (fixture_path.parent / "_rede_backups").exists()


def test_etag_rejects_stale_write(fixture_path, tmp_path):
    repository = ExcelDataStore(fixture_path, table="TabelaRede")
    client = create_app({"TESTING": True, "LOG_DIR": str(tmp_path / "logs")}, repository).test_client()
    version = client.get("/api/items/1").get_json()["data"]["_etag"]
    assert client.patch("/api/items/1", json={"STATUS": "Concluído"}, headers={"If-Match": version}).status_code == 200
    response = client.patch("/api/items/1", json={"STATUS": "Pendente"}, headers={"If-Match": version})
    assert response.status_code == 412
    assert repository.get_item("1")["STATUS"] == "Concluído"
    assert client.delete("/api/items/1", headers={"If-Match": version}).status_code == 412


def test_unsupported_object_blocks_without_saving(fixture_path):
    with zipfile.ZipFile(fixture_path, "a") as archive:
        archive.writestr("customXml/item1.xml", "<sample/>")
    original = fixture_path.read_bytes()
    with pytest.raises(ExcelSafetyError, match="complexos"):
        store(fixture_path).update("1", {"STATUS": "Concluído"})
    assert fixture_path.read_bytes() == original


def test_read_only_mode(fixture_path):
    adapter = store(fixture_path, read_only=True)
    assert len(adapter.read_rows()) == 2
    with pytest.raises(ExcelSafetyError, match="somente leitura"):
        adapter.update("1", {"STATUS": "Concluído"})


def test_full_table_expands_same_table_and_keeps_formatting(fixture_path):
    adapter = store(fixture_path)
    for i in range(3):
        adapter.create({"IDCLIENTE": f"NEW-{i}", "CLIENTE": "Teste"})

    before = load_workbook(fixture_path, data_only=False)
    old_table = before["Dados"].tables["TabelaRede"]
    identity = (old_table.name, old_table.id, str(old_table.tableStyleInfo))
    source_style = copy(before["Dados"]["A6"]._style)
    validation_before = str(before["Dados"].data_validations)
    source_height = before["Dados"].row_dimensions[6].height
    before.close()
    with zipfile.ZipFile(fixture_path) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml")

    created = adapter.create({"IDCLIENTE": "NO-ROOM", "CLIENTE": "Teste"})
    assert adapter.get(created["id"])["IDCLIENTE"] == "NO-ROOM"

    check = load_workbook(fixture_path, data_only=False)
    table = check["Dados"].tables["TabelaRede"]
    assert (table.name, table.id, str(table.tableStyleInfo)) == identity
    assert table.ref == "A1:W7"
    assert table.autoFilter is None or table.autoFilter.ref == "A1:W7"
    assert check["Dados"]["A7"]._style == source_style
    assert check["Dados"]["W7"].value == '=IF(A7="",0,1)'
    assert check["Dados"].row_dimensions[7].height == source_height
    assert "O7" in str(check["Dados"].data_validations)
    assert str(check["Dados"].data_validations) != validation_before
    assert any(
        "R7" in str(conditional)
        for conditional in check["Dados"].conditional_formatting._cf_rules
    )
    assert check.sheetnames == ["Dados", "Resumo"]
    assert check["Resumo"]["D2"].value == '=COUNTA(Dados!A2:A6)'
    check.close()
    with zipfile.ZipFile(fixture_path) as archive:
        assert archive.read("xl/charts/chart1.xml") == chart_xml


def test_expanded_table_keeps_api_crud_and_dashboard(fixture_path, tmp_path):
    adapter = store(fixture_path)
    for i in range(3):
        adapter.create({"IDCLIENTE": f"NEW-{i}", "CLIENTE": "Teste"})
    repository = ExcelDataStore(fixture_path, table="TabelaRede")
    client = create_app({"TESTING": True, "LOG_DIR": str(tmp_path / "logs")}, repository).test_client()

    response = client.post("/api/items", json={
        "IDCLIENTE": "API-NEW",
        "CLIENTE": "Cliente API",
        "STATUS": "Pendente",
        "CIDADE": "Santos",
    })
    assert response.status_code == 201
    created = response.get_json()["data"]
    assert client.get("/api/dashboard").get_json()["data"]["kpis"]["total"] == 6

    response = client.patch(
        f"/api/items/{created['id']}",
        json={"STATUS": "Concluído"},
        headers={"If-Match": created["_etag"]},
    )
    assert response.status_code == 200
    updated = response.get_json()["data"]
    assert updated["STATUS"] == "Concluído"
    assert client.get("/api/dashboard?status=Concluído").get_json()["data"]["kpis"]["total"] == 1

    assert client.delete(
        f"/api/items/{created['id']}", headers={"If-Match": updated["_etag"]}
    ).status_code == 200
    assert client.get("/api/dashboard").get_json()["data"]["kpis"]["total"] == 5
    check = load_workbook(fixture_path)
    assert check["Dados"].tables["TabelaRede"].ref == "A1:W7"
    assert check["Dados"]["A7"].value is None
    assert check["Dados"]["A7"].has_style
    check.close()


def test_table_expansion_never_overwrites_content_below(fixture_path):
    adapter = store(fixture_path)
    for i in range(3):
        adapter.create({"IDCLIENTE": f"NEW-{i}", "CLIENTE": "Teste"})
    wb = load_workbook(fixture_path)
    wb["Dados"]["A7"] = "CONTEUDO-EXTERNO"
    wb.save(fixture_path)
    wb.close()
    original = fixture_path.read_bytes()
    with pytest.raises(ExcelSafetyError, match="conteúdo logo abaixo"):
        adapter.create({"IDCLIENTE": "BLOCKED", "CLIENTE": "Teste"})
    assert fixture_path.read_bytes() == original


def test_replace_failure_leaves_original_intact(fixture_path, monkeypatch):
    import excel_io
    original = fixture_path.read_bytes()
    def fail(*args):
        raise PermissionError("locked")
    monkeypatch.setattr(excel_io.os, "replace", fail)
    monkeypatch.setattr(excel_io.time, "sleep", lambda delay: None)
    with pytest.raises(ExcelSafetyError, match="bloqueado"):
        store(fixture_path).update("1", {"STATUS": "Concluído"})
    assert fixture_path.read_bytes() == original
    assert not list(fixture_path.parent.glob(".rede-stage-*"))


def test_transient_replace_error_retries(fixture_path, monkeypatch):
    import excel_io
    real_replace, calls = excel_io.os.replace, []
    def retry_then_replace(*args):
        calls.append(1)
        if len(calls) < 3:
            raise PermissionError("locked")
        return real_replace(*args)
    monkeypatch.setattr(excel_io.os, "replace", retry_then_replace)
    monkeypatch.setattr(excel_io.time, "sleep", lambda delay: None)
    assert store(fixture_path).update("1", {"STATUS": "Concluído"})["STATUS"] == "Concluído"
    assert len(calls) == 3


def test_external_change_detected_before_commit(fixture_path, monkeypatch):
    import excel_safe
    original = fixture_path.read_bytes()
    external_bytes = original + b"externally modified"
    real_guard = excel_safe.guard_roundtrip
    def alter_then_guard(*args):
        real_guard(*args)
        fixture_path.write_bytes(external_bytes)
    monkeypatch.setattr(excel_safe, "guard_roundtrip", alter_then_guard)
    with pytest.raises(ExcelSafetyError, match="alterado"):
        store(fixture_path).update("1", {"STATUS": "Concluído"})
    assert fixture_path.read_bytes() == external_bytes


def _process_patch(path, changes):
    return store(path).update("1", changes)["id"]


def test_two_processes_no_lost_update(fixture_path):
    with ProcessPoolExecutor(max_workers=2, mp_context=multiprocessing.get_context("spawn")) as pool:
        futures = [pool.submit(_process_patch, str(fixture_path), values) for values in (
            {"STATUS": "Concluído"}, {"CIDADE": "Santos"})]
        assert [future.result(timeout=30) for future in futures] == ["1", "1"]
    result = store(fixture_path).get("1")
    assert result["STATUS"] == "Concluído" and result["CIDADE"] == "Santos"


def test_lock_timeout(fixture_path):
    with local_lock(fixture_path):
        with pytest.raises(ExcelSafetyError, match="em uso"):
            with local_lock(fixture_path, timeout=0.1):
                pass


def test_dates_keep_existing_number_format(fixture_path):
    from datetime import datetime
    wb = load_workbook(fixture_path)
    wb["Dados"]["L2"].number_format = "dd/mm/yyyy"
    wb.save(fixture_path)
    wb.close()
    store(fixture_path).update("1", {"DATAAGENDAMENTO": "2026-08-29"})
    wb = load_workbook(fixture_path)
    assert isinstance(wb["Dados"]["L2"].value, datetime)
    assert wb["Dados"]["L2"].number_format == "dd/mm/yyyy"
    wb.close()


def test_plain_sheet_append_extends_matching_autofilter(tmp_path):
    path = tmp_path / "plain.xlsx"
    wb = Workbook()
    wb.active.title = "Dados"
    wb.active.append(["_ITEM_ID", *FIELDS])
    wb.active.append(["1", "CLIENT-1", "Cliente 1"])
    wb.active.auto_filter.ref = "A1:V2"
    wb.save(path)
    wb.close()
    adapter = SafeExcel(path, "Dados", FIELDS)
    adapter.create({"IDCLIENTE": "CLIENT-2", "CLIENTE": "Cliente 2"})
    check = load_workbook(path)
    assert check["Dados"].auto_filter.ref == "A1:V3"
    check.close()
